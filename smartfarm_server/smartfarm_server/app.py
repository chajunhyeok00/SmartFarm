import os
from openpyxl import Workbook, load_workbook
from flask import Flask, request, jsonify, render_template_string, redirect, url_for
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user
from datetime import datetime

app = Flask(__name__)
app.secret_key = "your_secret_key"

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"

# 예시 사용자
users = {"admin": {"password": "1234"}}

class User(UserMixin):
    def __init__(self, username):
        self.id = username

@login_manager.user_loader
def load_user(username):
    if username in users:
        return User(username)
    return None

# ✅ 여러 식물의 센서 데이터를 plant_id 기준으로 저장
plant_data = {}

# HTML 템플릿 (로그인 화면)
login_html = """
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>SmartFarm 로그인</title>
    <style>
        body {
            margin: 0;
            padding: 0;
            font-family: Arial;
            background-color: #f0f4f7;
            height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            flex-direction: column;
            position: relative;
        }
        .login-box {
            background-color: #ffffff;
            padding: 30px;
            border-radius: 12px;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
            width: 300px;
            text-align: center;
        }
        h1 {
            margin-bottom: 40px;
            font-size: 28px;
            color: #2e7d32;
        }
        input[type="text"], input[type="password"] {
            width: 90%;
            padding: 10px;
            margin: 8px 0;
            border: 1px solid #ccc;
            border-radius: 6px;
        }
        input[type="submit"] {
            margin-top: 15px;
            padding: 10px 20px;
            background-color: #2e7d32;
            border: none;
            color: white;
            border-radius: 6px;
            cursor: pointer;
            font-size: 14px;
        }
        input[type="submit"]:hover {
            background-color: #1b5e20;
        }
        .alert {
            position: absolute;
            top: 20px;
            right: 20px;
            background-color: #ffe0e0;
            color: #c00;
            padding: 10px 15px;
            border-radius: 8px;
            box-shadow: 0 0 5px rgba(0,0,0,0.1);
            display: flex;
            align-items: center;
        }
        .alert .close {
            margin-left: 10px;
            cursor: pointer;
            font-weight: bold;
            color: #900;
        }
    </style>
</head>
<body>
    <h1>🌿 SmartFarm</h1>

    {% if error %}
    <div class="alert" id="alertBox">
        {{ error }}
        <span class="close" onclick="document.getElementById('alertBox').style.display='none'">✖</span>
    </div>
    {% endif %}

    <div class="login-box">
        <form method="post">
            <div>
                <input type="text" name="username" placeholder="아이디" required>
            </div>
            <div>
                <input type="password" name="password" placeholder="비밀번호" required>
            </div>
            <div>
                <input type="submit" value="로그인">
            </div>
        </form>
    </div>
</body>
</html>
"""

EXCEL_FILE = "sensor_log.xlsx"

def save_to_excel(data):
    plant_id = data.get("plant_id")
    if not plant_id:
        return  # plant_id 없으면 무시

    # 파일이 없으면 생성
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = plant_id  # 첫 시트 이름을 plant_id로 설정
        headers = ["timestamp"] + [k for k in data.keys() if k not in ("plant_id", "timestamp")]
        ws.append(headers)
        wb.save(EXCEL_FILE)

    # 엑셀 파일 열기
    wb = load_workbook(EXCEL_FILE)

    # 시트가 없으면 생성
    if plant_id not in wb.sheetnames:
        ws = wb.create_sheet(title=plant_id)
        headers = ["timestamp"] + [k for k in data.keys() if k not in ("plant_id", "timestamp")]
        ws.append(headers)
    else:
        ws = wb[plant_id]

    # 데이터 행 구성
    row = [data.get("timestamp")]
    for key in ws[1][1:]:  # 첫 번째 행의 두 번째 열부터 확인
        row.append(data.get(key.value, ""))

    ws.append(row)
    wb.save(EXCEL_FILE)


@app.route("/sensor", methods=["POST"])
def receive_sensor_data():
    data = request.get_json()
    if not data or "plant_id" not in data:
        return jsonify({"status": "fail", "message": "No plant_id or JSON received"}), 400

    data["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    plant_id = data["plant_id"]

    print(f"📡 [{plant_id}] 센서 데이터 수신:", data)

    if plant_id not in plant_data:
        plant_data[plant_id] = []
    plant_data[plant_id].append(data)

    save_to_excel(data)   # 여기서 엑셀 저장 함수 호출!

    return jsonify({"status": "ok"}), 200

@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        if username in users and users[username]["password"] == password:
            login_user(User(username))
            return redirect(url_for("index"))
        error = "❌ 로그인 실패! 아이디나 비밀번호를 확인하세요."

    return render_template_string(login_html, error=error)

@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))

@app.route("/")
@login_required
def index():
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>SmartFarm 작물 목록</title>
        <style>
            body { font-family: Arial; margin: 30px; background-color: #f1f8f5; }
            h2 { color: #2e7d32; }
            .logout-btn {
                position: absolute;
                top: 20px;
                right: 30px;
            }
            .logout-btn a {
                text-decoration: none;
                color: #2e7d32;
                font-weight: bold;
            }
            .plant-buttons {
                display: flex;
                flex-wrap: wrap;
                gap: 12px;
                margin-top: 20px;
            }
            .plant-button {
                padding: 12px 20px;
                background-color: #a5d6a7;
                color: #1b5e20;
                text-decoration: none;
                border-radius: 8px;
                font-weight: bold;
            }
            .plant-button:hover {
                background-color: #81c784;
            }
        </style>
    </head>
    <body>
        <div class="logout-btn">
            <a href="/logout">🔓 로그아웃</a>
        </div>
        <h2>🌿 SmartFarm 작물 선택</h2>
        <div class="plant-buttons">
            <a class="plant-button" href="/plant/basil">basil</a>
            <a class="plant-button" href="/plant/mint">mint</a>
            <a class="plant-button" href="/plant/tomato">tomato</a>
        </div>
    </body>
    </html>
    """
    return render_template_string(html)

@app.route("/plant/<plant_id>")
@login_required
def show_plant(plant_id):
    data = plant_data.get(plant_id, [])
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset=\"UTF-8\">
        <title>{{ plant_id }} 센서 데이터</title>
        <script src=\"https://cdn.jsdelivr.net/npm/chart.js\"></script>
        <style>
            body { font-family: Arial; margin: 30px; background-color: #f9f9f9; }
            h2 { color: #2e7d32; }
            .button-bar { margin-bottom: 20px; }
            .button-bar a {
                padding: 10px 15px;
                background-color: #81c784;
                color: white;
                text-decoration: none;
                border-radius: 8px;
                font-weight: bold;
                margin-right: 10px;
            }
            .box-container {
                display: flex;
                gap: 15px;
                flex-wrap: wrap;
            }
            .data-box {
                background-color: #e8f5e9;
                border: 1px solid #ccc;
                border-radius: 8px;
                padding: 16px;
                width: 250px;
                box-shadow: 2px 2px 6px rgba(0,0,0,0.1);
            }
            .data-box h3 {
                margin-top: 0;
                font-size: 18px;
                color: #2e7d32;
            }
            .data-item {
                margin: 5px 0;
                font-size: 14px;
            }
        </style>
    </head>
<body>
    <a class="back-link" href="/">← 작물 목록으로 돌아가기</a>
    <h2>🌱 {{ plant_id }} 센서 데이터</h2>

    <div class="button-bar">
        <a href="/plant/{{ plant_id }}/chart">📊 그래프로 보기</a>
    </div>

    <div class="box-container" id="data-container">
        <!-- JS로 데이터 로딩 -->
    </div>

    <script>
    async function fetchSensorData() {
        const res = await fetch('/api/plant/{{ plant_id }}/data');
        const data = await res.json();

        const container = document.getElementById('data-container');
        container.innerHTML = '';

        if (data.length === 0) {
            container.innerHTML = '<p>이 식물에 대한 센서 데이터가 아직 없습니다.</p>';
            return;
        }

        data.forEach(entry => {
            const box = document.createElement('div');
            box.className = 'data-box';
            box.innerHTML = `<h3>📅 ${entry.timestamp || '시간 없음'}</h3>`;

            for (const [key, value] of Object.entries(entry)) {
                if (key !== 'timestamp' && key !== 'plant_id') {
                    box.innerHTML += `<div class='data-item'><strong>${key}:</strong> ${value}</div>`;
                }
            }

            container.appendChild(box);
        });
    }

    fetchSensorData();
    setInterval(fetchSensorData, 30000);
    </script>
</body>

    </html>
    """
    return render_template_string(html, plant_id=plant_id, data=data)

@app.route("/plant/<plant_id>/chart")
@login_required
def show_plant_chart(plant_id):
    html = """
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset=\"UTF-8\">
        <title>{{ plant_id }} 센서 그래프</title>
        <script src=\"https://cdn.jsdelivr.net/npm/chart.js\"></script>
        <style>
            body { font-family: Arial; margin: 30px; background-color: #f9f9f9; }
            h2 { color: #2e7d32; }
            canvas { margin-top: 20px; background-color: white; padding: 10px; border-radius: 10px; box-shadow: 0 0 5px rgba(0,0,0,0.1); }
        </style>
    </head>
    <body>
        <a href="/plant/{{ plant_id }}">← 센서 데이터로 돌아가기</a>
        <h2>🌡️ {{ plant_id }} 온도/습도 그래프</h2>
        <canvas id="sensorChart" width="600" height="300"></canvas>

        <script>
let myChart = null;

async function drawChart() {
    const res = await fetch('/api/plant/{{ plant_id }}/data');
    const data = await res.json();

    const labels = data.map(d => d.timestamp);
    const temp = data.map(d => d.temperature);
    const humid = data.map(d => d.humidity);

    const ctx = document.getElementById('sensorChart').getContext('2d');

    // 기존 차트가 있다면 제거
    if (myChart) {
        myChart.destroy();
    }

    myChart = new Chart(ctx, {
        type: 'line',
        data: {
            labels: labels,
            datasets: [
                {
                    label: '온도 (°C)',
                    data: temp,
                    borderColor: 'rgba(255, 99, 132, 1)',
                    fill: false
                },
                {
                    label: '습도 (%)',
                    data: humid,
                    borderColor: 'rgba(54, 162, 235, 1)',
                    fill: false
                }
            ]
        },
        options: {
            responsive: true,
            scales: {
                x: { title: { display: true, text: '시간' } },
                y: { title: { display: true, text: '값' } }
            }
        }
    });
}

drawChart(); // 첫 실행
setInterval(drawChart, 30000); // 30초마다 자동 새로고침
</script>

    </body>
    </html>
    """
    return render_template_string(html, plant_id=plant_id)

# AJAX 요청에 대한 JSON 데이터 반환 API
@app.route("/api/plant/<plant_id>/data")
@login_required
def api_plant_data(plant_id):
    data = plant_data.get(plant_id, [])
    return jsonify(data)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
